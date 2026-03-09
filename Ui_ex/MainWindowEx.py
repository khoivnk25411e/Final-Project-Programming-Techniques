from PyQt6.QtWidgets import (QMessageBox, QTableWidgetItem, QHeaderView,
                             QDialog, QVBoxLayout, QLabel, QMainWindow, QProgressDialog)
from PyQt6.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt6 import QtGui

# --- THƯ VIỆN GỬI EMAIL ---
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import qrcode
from io import BytesIO

# --------------------------

try:
    from matplotlib import pyplot as plt
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.backends.backend_qtagg import NavigationToolbar2QT as NavigationToolbar
    import matplotlib

    matplotlib.use("QtAgg")
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
from ui.MainWindow import Ui_MainWindow
from Ui_ex.EventDialogEx import EventDialogEx
from Ui_ex.AttendeeDialogEx import AttendeeDialogEx
from Ui_ex.RegistrationDialogEx import RegistrationDialogEx
from Ui_ex.UserDialogEx import UserDialogEx
from Ui_ex.ChangePasswordDialogEx import ChangePasswordDialogEx
from models.events import Events
from models.event import Event
from models.attendees import Attendees
from models.attendee import Attendee
from models.registrations import Registrations
from models.registration import Registration
from models.users import Users
from models.user import User
from datetime import datetime
import uuid

try:
    import qrcode
    from PyQt6.QtGui import QPixmap, QImage
    from io import BytesIO

    QR_AVAILABLE = True
except ImportError:
    QR_AVAILABLE = False


# =====================================================================
# THREAD GỬI EMAIL NGẦM (GIÚP APP KHÔNG BỊ ĐƠ KHI GỬI NHIỀU EMAIL)
# =====================================================================
class EmailSenderThread(QThread):
    progress = pyqtSignal(int, str)  # Tín hiệu cập nhật thanh Progress (tiến độ, câu thông báo)
    finished_task = pyqtSignal(int, int)  # Tín hiệu khi hoàn thành (số thành công, số thất bại)

    def __init__(self, email_data_list, event_name):
        super().__init__()
        self.email_data_list = email_data_list
        self.event_name = event_name

        # 🔴 BẠN HÃY THAY EMAIL VÀ APP PASSWORD CỦA BẠN VÀO ĐÂY NHÉ:
        self.SENDER_EMAIL = "nguyenthanhdangkhoa9h@gmail.com"
        self.APP_PASSWORD = "rqwd zbcn fszy qeon"

    def run(self):
        success = 0
        failed = 0
        total = len(self.email_data_list)

        try:
            # Kết nối tới server Gmail
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(self.SENDER_EMAIL, self.APP_PASSWORD)

            for i, data in enumerate(self.email_data_list):
                self.progress.emit(i + 1, f"Sending Email to {data['name']}...")

                try:
                    # 1. Tạo form Email
                    msg = MIMEMultipart()
                    msg['From'] = self.SENDER_EMAIL
                    msg['To'] = data['email']
                    msg['Subject'] = f"🎟 Your Ticket & QR Code - {self.event_name}"

                    # 2. Nội dung Email
                    body = f"""
                    <h2>Hello {data['name']},</h2>
                    <p>You have successfully registered for the event: <b>{self.event_name}</b>.</p>
                    <p>Your Registration Code is: <b style="color: blue; font-size: 18px;">{data['code']}</b></p>
                    <p>Please find your QR code attached to this email. Show it at the check-in desk.</p>
                    <br><p>Best regards,<br>Event Management Team</p>
                    """
                    msg.attach(MIMEText(body, 'html'))

                    # 3. Tạo hình QR Code ngầm định trong bộ nhớ ảo
                    qr = qrcode.QRCode(version=1, box_size=10, border=5)
                    qr.add_data(data['code'])
                    qr.make(fit=True)
                    img = qr.make_image(fill_color="black", back_color="white")

                    buf = BytesIO()
                    img.save(buf, format="PNG")
                    image_data = buf.getvalue()

                    # 4. Đính kèm QR Code vào email
                    image = MIMEImage(image_data, name=f"QR_{data['code']}.png")
                    msg.attach(image)

                    # 5. Thực hiện gửi
                    server.send_message(msg)
                    success += 1
                except Exception as e:
                    print(f"Error sending to {data['email']}: {e}")
                    failed += 1

            server.quit()
        except Exception as e:
            print(f"SMTP Server error: {e}")
            failed = total

        self.finished_task.emit(success, failed)


# =====================================================================


def _msg(parent, kind, title, text):
    """QMessageBox với style rõ ràng — tránh chữ trùng màu nền trên Windows."""
    box = QMessageBox(parent)
    box.setWindowTitle(title)
    box.setText(text)
    if kind == "info":
        box.setIcon(QMessageBox.Icon.Information)
    elif kind == "warn":
        box.setIcon(QMessageBox.Icon.Warning)
    elif kind == "err":
        box.setIcon(QMessageBox.Icon.Critical)
    box.setStyleSheet("""
        QMessageBox        { background-color: white; }
        QMessageBox QLabel { color: #111827; font-size: 13px; min-width: 260px; }
        QMessageBox QPushButton {
            background: #3b82f6; color: white; border: none;
            padding: 7px 22px; border-radius: 5px;
            font-size: 12px; min-width: 80px;
        }
        QMessageBox QPushButton:hover { background: #2563eb; }
    """)
    box.exec()


class MainWindowEx(Ui_MainWindow):
    def setupUi(self, MainWindow):
        super().setupUi(MainWindow)
        self.MainWindow = MainWindow
        self._apply_role_ui()
        self.setupSignalAndSlot()
        self.load_initial_data()
        self.apply_stylesheet()

    def showWindow(self):
        self.MainWindow.show()

    def _apply_role_ui(self):
        user = getattr(self, 'login_user', None)
        if user is None:
            return

        is_admin = (user.Role == "admin")

        self.lblLoginUser.setText(f"👤  {user.FullName}  ({user.UserName})")
        if is_admin:
            self.lblRoleBadge.setText("ADMIN")
            self.lblRoleBadge.setStyleSheet(
                "QLabel{background:#ef4444;color:white;border-radius:5px;"
                "font-size:11px;font-weight:bold;padding:0 8px;}"
            )
        else:
            self.lblRoleBadge.setText("STAFF")
            self.lblRoleBadge.setStyleSheet(
                "QLabel{background:#3b82f6;color:white;border-radius:5px;"
                "font-size:11px;font-weight:bold;padding:0 8px;}"
            )

        idx = self.tabWidget.indexOf(self.userMgmtTab)
        if not is_admin and idx >= 0:
            self.tabWidget.removeTab(idx)

        if not is_admin:
            self.btnAddEvent.setEnabled(False)
            self.btnEditEvent.setEnabled(False)
            self.btnDeleteEvent.setEnabled(False)

            self.btnAddAttendee.setEnabled(False)
            self.btnEditAttendee.setEnabled(False)
            self.btnDeleteAttendee.setEnabled(False)

            self.btnRegisterAttendee.setEnabled(False)
            self.btnCancelRegistration.setEnabled(False)
            self.btnGenerateQR.setEnabled(False)

            self.btnCheckin.setEnabled(True)
            self.btnScanQR.setEnabled(True)
            self.checkinCode.setEnabled(True)
            self.checkinCode.setPlaceholderText("Enter check-in code or scan QR")

            self.btnRefreshEvent.setEnabled(True)
            self.btnRefreshAttendee.setEnabled(True)
            self.btnRefreshRegistration.setEnabled(True)
            self.btnRefreshCheckin.setEnabled(True)

    def setupSignalAndSlot(self):
        self.btnChangePassword.clicked.connect(self.change_password)
        self.btnLogout.clicked.connect(self.logout)

        self.eventSearch.textChanged.connect(self.search_events)
        self.btnAddEvent.clicked.connect(self.add_event)
        self.btnViewEvent.clicked.connect(self.view_event_details)
        self.btnEditEvent.clicked.connect(self.edit_event)
        self.btnDeleteEvent.clicked.connect(self.delete_event)
        self.btnRefreshDash.clicked.connect(self.load_dashboard)
        self.btnImportAttendee.clicked.connect(self.import_attendees_from_file)
        self.btnDownloadTemplate.clicked.connect(self.download_import_template)
        self.btnExportEvent.clicked.connect(self.export_events_csv)
        self.btnPrevEvent.clicked.connect(lambda: self._prev_page("eventTable"))
        self.btnNextEvent.clicked.connect(lambda: self._next_page("eventTable"))

        self.attendeeSearch.textChanged.connect(self.search_attendees)
        self.btnAddAttendee.clicked.connect(self.add_attendee)
        self.btnEditAttendee.clicked.connect(self.edit_attendee)
        self.btnDeleteAttendee.clicked.connect(self.delete_attendee)
        self.btnHistoryAttendee.clicked.connect(self.view_attendee_history)
        self.btnRefreshAttendee.clicked.connect(self.load_attendees)
        self.btnExportAttendee.clicked.connect(self.export_attendees_csv)
        self.btnPrevAttendee.clicked.connect(lambda: self._prev_page("attendeeTable"))
        self.btnNextAttendee.clicked.connect(lambda: self._next_page("attendeeTable"))

        self.eventCombo.currentIndexChanged.connect(self.load_registrations)
        self.btnRegisterAttendee.clicked.connect(self.register_attendee)
        self.btnGenerateQR.clicked.connect(self.generate_qr_code)
        self.btnCancelRegistration.clicked.connect(self.cancel_registration)
        self.btnRefreshRegistration.clicked.connect(self.load_registrations)
        self.btnExportRegistration.clicked.connect(self.export_registrations_csv)
        self.registrationSearch.textChanged.connect(self.search_registrations)
        self.btnPrevRegistration.clicked.connect(lambda: self._prev_page("registrationTable"))
        self.btnNextRegistration.clicked.connect(lambda: self._next_page("registrationTable"))

        self.checkinEventCombo.currentIndexChanged.connect(self.load_checkin_stats)
        self.btnCheckin.clicked.connect(self.perform_checkin)
        self.btnScanQR.clicked.connect(self.scan_qr_checkin)
        self.btnRefreshCheckin.clicked.connect(self.load_checkin_stats)
        self.btnExportCheckin.clicked.connect(self.export_checkin_csv)

        self.statsEventCombo.currentIndexChanged.connect(self.refresh_chart)
        self.btnStatsBar.clicked.connect(self.show_bar_chart)
        self.btnStatsLine.clicked.connect(self.show_line_chart)
        self.btnStatsPie.clicked.connect(self.show_pie_chart)

        user = getattr(self, 'login_user', None)
        if user and user.Role == "admin":
            self.btnAddUser.clicked.connect(self.add_user)
            self.btnEditUser.clicked.connect(self.edit_user)
            self.btnDeleteUser.clicked.connect(self.delete_user)
            self.btnResetUserPwd.clicked.connect(self.reset_user_password)
            self.btnRefreshUser.clicked.connect(self.load_users)
        else:
            self.btnRefreshDash.clicked.connect(self.load_dashboard)
            self.btnImportAttendee.clicked.connect(self.import_attendees_from_file)
            self.btnDownloadTemplate.clicked.connect(self.download_import_template)
            self.btnRefreshEvent.clicked.connect(self.load_events)
            self.btnRefreshAttendee.clicked.connect(self.load_attendees)
            self.btnRefreshRegistration.clicked.connect(self.load_registrations)
            self.btnRefreshCheckin.clicked.connect(self.load_checkin_stats)
            self.btnCheckin.clicked.connect(self.perform_checkin)
            self.btnScanQR.clicked.connect(self.scan_qr_checkin)
            self.btnExportEvent.clicked.connect(self.export_events_csv)
            self.btnExportAttendee.clicked.connect(self.export_attendees_csv)
            self.btnExportRegistration.clicked.connect(self.export_registrations_csv)
            self.btnExportCheckin.clicked.connect(self.export_checkin_csv)
            self.registrationSearch.textChanged.connect(self.search_registrations)

        all_tables = [self.eventTable, self.attendeeTable,
                      self.registrationTable, self.checkinTable]
        if user and user.Role == "admin":
            all_tables.append(self.userTable)

        for tbl in all_tables:
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            tbl.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

    def load_initial_data(self):
        self.load_dashboard()
        self.load_events()
        self.load_attendees()
        self.load_event_combo()
        self.load_checkin_event_combo()
        self.load_stats_event_combo()

        for tbl in [self.eventTable, self.attendeeTable,
                    self.registrationTable, self.checkinTable, self.userTable]:
            tbl.setAlternatingRowColors(True)

        from PyQt6.QtWidgets import QAbstractItemView
        self.attendeeTable.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.registrationTable.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.setup_chart()
        user = getattr(self, 'login_user', None)
        if user and user.Role == "admin":
            self.load_users()

    def change_password(self):
        user = getattr(self, 'login_user', None)
        if not user:
            return
        dlg = ChangePasswordDialogEx(self.MainWindow, current_user=user, skip_old_password=False)
        dlg.exec()

    def logout(self):
        reply = QMessageBox.question(
            self.MainWindow, "Logout",
            "Are you sure you want to log out?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.MainWindow.close()
            from Ui_ex.LoginWindowEx import LoginWindowEx
            self.login_gui = LoginWindowEx()
            self.login_gui.setupUi(QMainWindow())
            self.login_gui.showWindow()

    def load_users(self):
        users = Users()
        users.import_json("datasets/users.json")
        role_lbl = lambda r: "🔑 Admin" if r == "admin" else "🧑‍💼 Staff"
        rows_data = [
            [u.UserId, u.FullName, u.UserName, u.Email, role_lbl(u.Role)]
            for u in users.list
        ]
        self._fill_table(self.userTable, rows_data)
        self._pad_table(self.userTable, len(rows_data))

    def _users_db(self):
        u = Users()
        u.import_json("datasets/users.json")
        return u

    def _apply_user_data(self, user, data):
        user.FullName = data['full_name']
        user.UserName = data['username']
        user.Email = data['email']
        user.Role = data['role']
        user.SecurityQuestion = data['sec_question']
        user.SecurityAnswer = data['sec_answer']

    def add_user(self):
        dlg = UserDialogEx(self.MainWindow)
        if dlg.exec() != QDialog.DialogCode.Accepted: return
        data = dlg.get_data()
        users = self._users_db()
        if users.find_by_username(data['username']):
            _msg(self.MainWindow, "warn", "Error", "Username already exists!")
            return
        if users.find_by_email(data['email']):
            _msg(self.MainWindow, "warn", "Error", "Email already exists!")
            return
        new_user = User()
        new_user.UserId = "usr_" + str(uuid.uuid4())[:8]
        self._apply_user_data(new_user, data)
        new_user.Password = Users.hash_password(data['password'])
        users.add_item(new_user)
        users.export_json("datasets/users.json")
        _msg(self.MainWindow, "info", "Success", "New account created!")
        self.load_users()

    def edit_user(self):
        row = self.userTable.currentRow()
        if row < 0:
            _msg(self.MainWindow, "warn", "Error", "Please select an account to edit!")
            return
        user_id = self.userTable.item(row, 0).text()
        users = self._users_db()
        user = users.find_user(user_id)
        if not user: return
        dlg = UserDialogEx(self.MainWindow, user_data=user)
        if dlg.exec() != QDialog.DialogCode.Accepted: return
        data = dlg.get_data()
        un = users.find_by_username(data['username'])
        em = users.find_by_email(data['email'])
        if un and un.UserId != user_id:
            _msg(self.MainWindow, "warn", "Error", "Username already exists!")
            return
        if em and em.UserId != user_id:
            _msg(self.MainWindow, "warn", "Error", "Email already exists!")
            return
        self._apply_user_data(user, data)
        if data['password']: user.Password = Users.hash_password(data['password'])
        users.export_json("datasets/users.json")
        _msg(self.MainWindow, "info", "Success", "Account updated!")
        self.load_users()

    def delete_user(self):
        row = self.userTable.currentRow()
        if row < 0:
            _msg(self.MainWindow, "warn", "Error", "Please select an account to delete!")
            return
        user_id = self.userTable.item(row, 0).text()

        if getattr(self, 'login_user', None) and self.login_user.UserId == user_id:
            _msg(self.MainWindow, "warn", "Error", "You cannot delete the currently logged-in account!")
            return

        reply = QMessageBox.question(
            self.MainWindow, "Confirm", "Are you sure you want to delete this account?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            users = Users()
            users.import_json("datasets/users.json")
            users.delete_user(user_id)
            users.export_json("datasets/users.json")
            _msg(self.MainWindow, "info", "Success", "Account deleted successfully!")
            self.load_users()

    def reset_user_password(self):
        row = self.userTable.currentRow()
        if row < 0:
            _msg(self.MainWindow, "warn", "Error", "Please select an account!")
            return
        user_id = self.userTable.item(row, 0).text()
        users = Users()
        users.import_json("datasets/users.json")
        user = users.find_user(user_id)
        if user:
            dlg = ChangePasswordDialogEx(self.MainWindow, current_user=user, skip_old_password=True)
            dlg.exec()

    PAGE_SIZE = 10

    def _pad_table(self, table, actual_count):
        """Đặt số rows = PAGE_SIZE để lấp đầy khoảng trống, rows thừa để trống."""
        if actual_count < self.PAGE_SIZE:
            table.setRowCount(self.PAGE_SIZE)

    def _fill_table(self, table, rows_data):
        table.setRowCount(len(rows_data))
        for row, cols in enumerate(rows_data):
            for col, val in enumerate(cols):
                table.setItem(row, col, QTableWidgetItem(str(val) if val else ""))

    def search_events(self):
        keyword = self.eventSearch.text().strip()
        if not keyword:
            self.load_events()
            return
        events = Events()
        events.import_json("datasets/events.json")
        keyword_lower = keyword.lower()
        results = [
            e for e in events.list
            if keyword_lower in e.EventName.lower()
               or keyword_lower in e.Location.lower()
               or keyword_lower in (e.Description or "").lower()
        ]
        self._setup_pagination(
            self.eventTable, results,
            self._fill_event_table,
            self.lblPageEvent, self.btnPrevEvent, self.btnNextEvent
        )

    def _fill_event_table(self, event_list):
        self.eventTable.setRowCount(len(event_list))
        for row, event in enumerate(event_list):
            self.eventTable.setItem(row, 0, QTableWidgetItem(event.EventId))
            self.eventTable.setItem(row, 1, QTableWidgetItem(event.EventName))
            self.eventTable.setItem(row, 2, QTableWidgetItem(event.EventDate))
            self.eventTable.setItem(row, 3, QTableWidgetItem(event.EventTime))
            self.eventTable.setItem(row, 4, QTableWidgetItem(event.Location))
            self.eventTable.setItem(row, 5, QTableWidgetItem(event.Description or ""))
        self._pad_table(self.eventTable, len(event_list))

    def load_events(self):
        events = Events()
        events.import_json("datasets/events.json")
        self._setup_pagination(
            self.eventTable, events.list,
            self._fill_event_table,
            self.lblPageEvent, self.btnPrevEvent, self.btnNextEvent
        )

    def add_event(self):
        dlg = EventDialogEx(self.MainWindow)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            data = dlg.get_data()
            event = Event()
            event.EventId = "evt_" + str(uuid.uuid4())[:8]
            event.EventName = data['name']
            event.EventDate = data['date']
            event.EventTime = data['time']
            event.Location = data['location']
            event.Description = data['description']
            events = Events()
            events.import_json("datasets/events.json")
            events.add_item(event)
            events.export_json("datasets/events.json")
            _msg(self.MainWindow, "info", "Success", "New event added successfully!")
            self.load_events()
            self.load_event_combo()
            self.load_checkin_event_combo()

    def edit_event(self):
        row = self.eventTable.currentRow()
        if row < 0:
            _msg(self.MainWindow, "warn", "Error", "Please select an event to edit!")
            return
        event_id = self.eventTable.item(row, 0).text()
        events = Events()
        events.import_json("datasets/events.json")
        event = events.find_event(event_id)
        if event:
            dlg = EventDialogEx(self.MainWindow, event)
            if dlg.exec() == QDialog.DialogCode.Accepted:
                data = dlg.get_data()
                event.EventName = data['name']
                event.EventDate = data['date']
                event.EventTime = data['time']
                event.Location = data['location']
                event.Description = data['description']
                events.export_json("datasets/events.json")
                _msg(self.MainWindow, "info", "Success", "Event updated successfully!")
                self.load_events()
                self.load_event_combo()
                self.load_checkin_event_combo()

    def delete_event(self):
        row = self.eventTable.currentRow()
        if row < 0:
            _msg(self.MainWindow, "warn", "Error", "Please select an event to delete!")
            return
        reply = QMessageBox.question(
            self.MainWindow, "Confirm",
            "Are you sure you want to delete this event?\nAll related registrations will be deleted!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            event_id = self.eventTable.item(row, 0).text()
            events = Events()
            events.import_json("datasets/events.json")
            events.delete_event(event_id)
            events.export_json("datasets/events.json")
            regs = Registrations()
            regs.import_json("datasets/registrations.json")
            deleted = regs.delete_by_event(event_id)
            regs.export_json("datasets/registrations.json")
            _msg(self.MainWindow, "info", "Success", f"Event deleted! ({deleted} registrations removed)")
            self.load_events()
            self.load_event_combo()
            self.load_checkin_event_combo()

    def view_event_details(self):
        row = self.eventTable.currentRow()
        if row < 0:
            _msg(self.MainWindow, "warn", "Error", "Please select an event!")
            return
        event_id = self.eventTable.item(row, 0).text()
        events = Events()
        events.import_json("datasets/events.json")
        event = events.find_event(event_id)
        if event:
            regs = Registrations()
            regs.import_json("datasets/registrations.json")
            total_reg = regs.count_registered_by_event(event_id)
            total_checkin = regs.count_checkedin_by_event(event_id)
            details = (
                f"<h2>{event.EventName}</h2>"
                f"<p><b>📅 Date:</b> {event.EventDate}</p>"
                f"<p><b>🕐 Time:</b> {event.EventTime}</p>"
                f"<p><b>📍 Location:</b> {event.Location}</p>"
                f"<p><b>📝 Description:</b> {event.Description or 'N/A'}</p>"
                f"<hr><p><b>👥 Total Registered:</b> {total_reg}</p>"
                f"<p><b>✅ Checked-in:</b> {total_checkin}</p>"
            )
            _msg(self.MainWindow, "info", "Event Details", details)

    def _fill_attendee_table(self, attendee_list):
        self.attendeeTable.setRowCount(len(attendee_list))
        for row, a in enumerate(attendee_list):
            self.attendeeTable.setItem(row, 0, QTableWidgetItem(a.AttendeeId))
            self.attendeeTable.setItem(row, 1, QTableWidgetItem(a.Name))
            self.attendeeTable.setItem(row, 2, QTableWidgetItem(a.Email))
            self.attendeeTable.setItem(row, 3, QTableWidgetItem(a.Phone or ""))
            self.attendeeTable.setItem(row, 4, QTableWidgetItem(a.Organization or ""))
            self.attendeeTable.setItem(row, 5, QTableWidgetItem(a.Position or ""))
        self._pad_table(self.attendeeTable, len(attendee_list))

    def load_attendees(self):
        attendees = Attendees()
        attendees.import_json("datasets/attendees.json")
        self._setup_pagination(
            self.attendeeTable, attendees.list,
            self._fill_attendee_table,
            self.lblPageAttendee, self.btnPrevAttendee, self.btnNextAttendee
        )

    def search_attendees(self):
        keyword = self.attendeeSearch.text().strip()
        if not keyword:
            self.load_attendees()
            return
        attendees = Attendees()
        attendees.import_json("datasets/attendees.json")
        results = attendees.search_attendees(keyword)
        self._setup_pagination(
            self.attendeeTable, results,
            self._fill_attendee_table,
            self.lblPageAttendee, self.btnPrevAttendee, self.btnNextAttendee
        )

    def add_attendee(self):
        dlg = AttendeeDialogEx(self.MainWindow)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            data = dlg.get_data()
            attendees = Attendees()
            attendees.import_json("datasets/attendees.json")
            for a in attendees.list:
                if a.Email.lower() == data['email'].lower():
                    QMessageBox.warning(self.MainWindow, "Error", "Email already exists!")
                    return
            att = Attendee()
            att.AttendeeId = "att_" + str(uuid.uuid4())[:8]
            att.Name = data['name']
            att.Email = data['email']
            att.Phone = data['phone']
            att.Organization = data['organization']
            att.Position = data['position']
            attendees.add_item(att)
            attendees.export_json("datasets/attendees.json")
            _msg(self.MainWindow, "info", "Success", "Attendee added successfully!")
            self.load_attendees()

    def edit_attendee(self):
        row = self.attendeeTable.currentRow()
        if row < 0:
            _msg(self.MainWindow, "warn", "Error", "Please select an attendee to edit!")
            return
        att_id = self.attendeeTable.item(row, 0).text()
        attendees = Attendees()
        attendees.import_json("datasets/attendees.json")
        att = attendees.find_attendee(att_id)
        if att:
            dlg = AttendeeDialogEx(self.MainWindow, att)
            if dlg.exec() == QDialog.DialogCode.Accepted:
                data = dlg.get_data()
                for a in attendees.list:
                    if a.Email.lower() == data['email'].lower() and a.AttendeeId != att_id:
                        _msg(self.MainWindow, "warn", "Error", "Email already exists!")
                        return
                att.Name = data['name']
                att.Email = data['email']
                att.Phone = data['phone']
                att.Organization = data['organization']
                att.Position = data['position']
                attendees.export_json("datasets/attendees.json")
                _msg(self.MainWindow, "info", "Success", "Attendee updated successfully!")
                self.load_attendees()

    def delete_attendee(self):
        rows = list(set(idx.row() for idx in self.attendeeTable.selectedIndexes()))
        if not rows:
            _msg(self.MainWindow, "warn", "Error", "Please select at least one attendee!")
            return
        reply = QMessageBox.question(
            self.MainWindow, "Confirm",
            f"Delete {len(rows)} attendee(s)?\nAll related registrations will also be deleted!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            att_ids = [self.attendeeTable.item(r, 0).text() for r in rows]
            attendees = Attendees()
            attendees.import_json("datasets/attendees.json")
            regs = Registrations()
            regs.import_json("datasets/registrations.json")
            total_regs = 0
            for att_id in att_ids:
                attendees.delete_attendee(att_id)
                total_regs += regs.delete_by_attendee(att_id)
            attendees.export_json("datasets/attendees.json")
            regs.export_json("datasets/registrations.json")
            _msg(self.MainWindow, "info", "Success",
                 f"Deleted {len(att_ids)} attendee(s)! ({total_regs} registrations removed)")
            self.load_attendees()

    def search_registrations(self):
        keyword = self.registrationSearch.text().strip().lower()
        if not keyword:
            self.load_registrations()
            return
        event_id = self.eventCombo.currentData()
        regs = Registrations()
        regs.import_json("datasets/registrations.json")
        attendees = Attendees()
        attendees.import_json("datasets/attendees.json")
        all_regs = regs.get_registrations_by_event(event_id) if event_id else regs.list

        filtered = []
        for r in all_regs:
            att = attendees.find_attendee(r.AttendeeId)
            name = att.Name.lower() if att else ""
            email = att.Email.lower() if att else ""
            org = att.Organization.lower() if att else ""
            status = r.Status.lower()
            if keyword in name or keyword in email or keyword in org or keyword in status:
                filtered.append(r)

        self.registrationTable.setRowCount(len(filtered))
        for row, r in enumerate(filtered):
            att = attendees.find_attendee(r.AttendeeId)
            name = att.Name if att else "Unknown"
            email = att.Email if att else ""
            org = att.Organization if att else ""
            from PyQt6.QtWidgets import QTableWidgetItem
            from PyQt6.QtGui import QColor
            vals = [r.RegistrationId, name, email, org, r.RegistrationDate, r.Status]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                if r.Status == "Checked-in":
                    item.setForeground(QColor("#16a34a"))
                self.registrationTable.setItem(row, col, item)

    def load_dashboard(self):
        from datetime import datetime
        from PyQt6.QtWidgets import QTableWidgetItem
        from PyQt6.QtGui import QColor

        events = Events()
        events.import_json("datasets/events.json")
        attendees = Attendees()
        attendees.import_json("datasets/attendees.json")
        regs = Registrations()
        regs.import_json("datasets/registrations.json")

        today = datetime.now().strftime("%d/%m/%Y")

        def set_card_num(card_widget, num):
            for child in card_widget.findChildren(__import__('PyQt6.QtWidgets', fromlist=['QLabel']).QLabel):
                if child.objectName().endswith("Num"):
                    child.setText(str(num))

        set_card_num(self.cardTotalEvents, len(events.list))
        set_card_num(self.cardTotalAttendees, len(attendees.list))
        set_card_num(self.cardTotalRegs, len(regs.list))
        checkin_today = sum(1 for r in regs.list
                            if r.Status == "Checked-in" and
                            (r.CheckinTime or "").startswith(datetime.now().strftime("%Y-%m-%d")))
        set_card_num(self.cardTotalCheckin, checkin_today)

        def parse_date(d):
            try:
                parts = d.split('/')
                return datetime(int(parts[2]), int(parts[1]), int(parts[0]))
            except Exception:
                return datetime(2000, 1, 1)

        now = datetime.now()
        upcoming = sorted(
            [e for e in events.list if parse_date(e.EventDate) >= now],
            key=lambda e: parse_date(e.EventDate)
        )[:8]

        self.dashUpcomingTable.setRowCount(len(upcoming))
        for i, e in enumerate(upcoming):
            reg_count = regs.count_registered_by_event(e.EventId)
            for col, val in enumerate([e.EventName, e.EventDate, str(reg_count)]):
                item = QTableWidgetItem(val)
                if col == 1 and e.EventDate == today:
                    item.setForeground(QColor("#16a34a"))
                    item.setText(f"TODAY — {val}")
                self.dashUpcomingTable.setItem(i, col, item)

        recent = sorted(
            [r for r in regs.list if r.Status == "Checked-in" and r.CheckinTime],
            key=lambda r: r.CheckinTime or "",
            reverse=True
        )[:10]

        self.dashRecentTable.setRowCount(len(recent))
        for i, r in enumerate(recent):
            att = attendees.find_attendee(r.AttendeeId)
            evt = events.find_event(r.EventId)
            name = att.Name if att else "Unknown"
            evt_name = evt.EventName if evt else "Unknown"
            time_str = (r.CheckinTime or "")[-8:] if r.CheckinTime else ""
            for col, val in enumerate([name, evt_name, time_str]):
                item = QTableWidgetItem(val)
                item.setForeground(QColor("#16a34a"))
                self.dashRecentTable.setItem(i, col, item)

    def download_import_template(self):
        """Tạo và tải file mẫu CSV/Excel để import attendees"""
        from PyQt6.QtWidgets import QFileDialog
        import csv

        path, selected_filter = QFileDialog.getSaveFileName(
            self.MainWindow,
            "Save Import Template",
            "attendees_template",
            "CSV File (*.csv);;Excel File (*.xlsx)"
        )
        if not path:
            return

        # Dữ liệu mẫu
        headers = ["name", "email", "phone", "organization", "position"]
        sample_rows = [
            ["Nguyễn Văn A", "nguyenvana@email.com", "0901234567", "FPT Software", "Developer"],
            ["Trần Thị B",   "tranthib@email.com",   "0912345678", "VNG Corporation", "Manager"],
            ["Lê Văn C",     "levanc@email.com",      "0923456789", "HCMUT", "Student"],
        ]

        try:
            if path.endswith('.xlsx'):
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Attendees Template"

                    # Header style: nền xanh, chữ trắng, bold
                    header_fill = PatternFill("solid", fgColor="3B82F6")
                    header_font = Font(bold=True, color="FFFFFF", size=11)
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                        ws.column_dimensions[cell.column_letter].width = 22

                    # Sample rows: nền vàng nhạt
                    sample_fill = PatternFill("solid", fgColor="FEF9C3")
                    for r_idx, row in enumerate(sample_rows, 2):
                        for c_idx, val in enumerate(row, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=val)
                            cell.fill = sample_fill

                    # Ghi chú ở dòng cuối
                    note_row = len(sample_rows) + 3
                    ws.cell(row=note_row, column=1,
                            value="※ Xóa các dòng mẫu trên và điền dữ liệu thật vào. Giữ nguyên dòng header.")
                    ws.cell(row=note_row, column=1).font = Font(italic=True, color="6B7280")

                    if not path.endswith('.xlsx'):
                        path += '.xlsx'
                    wb.save(path)
                except ImportError:
                    _msg(self.MainWindow, "warn", "Missing Library",
                         "openpyxl not installed!\nSaving as CSV instead.\n\nInstall: pip install openpyxl")
                    path = path.replace('.xlsx', '.csv')
                    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                        writer = csv.writer(f)
                        writer.writerow(headers)
                        writer.writerows(sample_rows)
            else:
                # CSV
                if not path.endswith('.csv'):
                    path += '.csv'
                with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(sample_rows)

            _msg(self.MainWindow, "info", "✅ Template Saved",
                 f"Template saved!\n\n📋 Columns: name, email, phone, organization, position\n\n"
                 f"Fill in your data and use '📂 Import Excel/CSV' to import.\n\n📁 {path}")

        except Exception as e:
            _msg(self.MainWindow, "err", "Error", f"Cannot save template:\n{e}")

    def import_attendees_from_file(self):
        import csv
        from PyQt6.QtWidgets import QFileDialog

        path, _ = QFileDialog.getOpenFileName(
            self.MainWindow, "Import Attendees",
            "", "Excel/CSV Files (*.xlsx *.xls *.csv);;All Files (*)"
        )
        if not path:
            return

        rows = []
        errors = []

        try:
            if path.lower().endswith('.csv'):
                with open(path, encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for i, row in enumerate(reader, 2):
                        rows.append((i, row))
            else:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(path)
                    ws = wb.active
                    headers = [str(c.value or "").strip().lower() for c in ws[1]]
                    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                        rows.append((i, dict(zip(headers, row))))
                except ImportError:
                    _msg(self.MainWindow, "warn", "Missing Library",
                         "openpyxl not installed!\nInstall with: pip install openpyxl\n\nOr use CSV format instead.")
                    return
        except Exception as e:
            QMessageBox.critical(self.MainWindow, "Error", f"Cannot read file:\n{e}")
            return

        def get_val(row_dict, *keys):
            for k in keys:
                for dk in row_dict:
                    if str(dk).strip().lower() == k.lower():
                        v = row_dict[dk]
                        return str(v).strip() if v is not None else ""
            return ""

        attendees = Attendees()
        attendees.import_json("datasets/attendees.json")

        import uuid, re
        added = skipped = failed = 0

        for line_num, row in rows:
            name = get_val(row, 'name', 'full name', 'fullname', 'họ tên', 'ten')
            email = get_val(row, 'email', 'e-mail', 'email address')
            phone = get_val(row, 'phone', 'phone number', 'điện thoại', 'sdt')
            org = get_val(row, 'organization', 'company', 'tổ chức', 'don vi')
            pos = get_val(row, 'position', 'job title', 'chức vụ', 'chuc vu')

            if not name or not email:
                errors.append(f"Row {line_num}: Missing name or email")
                failed += 1
                continue

            if not re.match(r'^[\w.+-]+@[\w-]+\.[\w.]+$', email):
                errors.append(f"Row {line_num}: Invalid email '{email}'")
                failed += 1
                continue

            if attendees.is_email_taken(email):
                skipped += 1
                continue

            att = type('Attendee', (), {})()
            att.AttendeeId = "att_" + str(uuid.uuid4())[:8]
            att.Name = name
            att.Email = email
            att.Phone = phone
            att.Organization = org
            att.Position = pos
            attendees.add_item(att)
            added += 1

        attendees.export_json("datasets/attendees.json")
        self.load_attendees()
        self.load_dashboard()

        msg = f"Import complete!\n\nAdded: {added}\nSkipped (duplicate email): {skipped}\nFailed: {failed}"
        if errors:
            msg += "\n\nErrors (first 5):\n" + "\n".join(errors[:5])
        _msg(self.MainWindow, "info", "Import Result", msg)

    def view_attendee_history(self):
        row = self.attendeeTable.currentRow()
        if row < 0:
            QMessageBox.warning(self.MainWindow, "Error", "Please select an attendee!")
            return

        att_id = self.attendeeTable.item(row, 0).text()
        att_name = self.attendeeTable.item(row, 1).text()

        regs = Registrations()
        regs.import_json("datasets/registrations.json")
        events = Events()
        events.import_json("datasets/events.json")

        att_regs = [r for r in regs.list if r.AttendeeId == att_id]

        from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QLabel,
                                     QTableWidget, QTableWidgetItem, QHeaderView)
        from PyQt6.QtGui import QColor

        dlg = QDialog(self.MainWindow)
        dlg.setWindowTitle(f"📜 History: {att_name}")
        dlg.setMinimumSize(700, 400)
        dlg.setStyleSheet("background: white;")

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        lbl = QLabel(f"<b>{att_name}</b> — {len(att_regs)} event(s) registered")
        lbl.setStyleSheet("font-size: 14px; color: #111827; padding-bottom: 4px;")
        layout.addWidget(lbl)

        table = QTableWidget(len(att_regs), 5)
        table.setHorizontalHeaderLabels(["Event Name", "Date", "Location", "Reg. Date", "Status"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.setStyleSheet("background: white; alternate-background-color: #f8fafc; gridline-color: #e2e8f0;")
        table.horizontalHeader().setStyleSheet(
            "QHeaderView::section { background: #f1f5f9; color: #374151; font-weight: bold; padding: 6px; border-bottom: 2px solid #e2e8f0; }"
        )

        for r_idx, reg in enumerate(att_regs):
            evt = events.find_event(reg.EventId)
            evt_name = evt.EventName if evt else "Unknown"
            evt_date = evt.EventDate if evt else ""
            evt_loc = evt.Location if evt else ""

            vals = [evt_name, evt_date, evt_loc, reg.RegistrationDate, reg.Status]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                if reg.Status == "Checked-in":
                    item.setForeground(QColor("#16a34a"))
                table.setItem(r_idx, c, item)

        layout.addWidget(table)

        if not att_regs:
            no_data = QLabel("This attendee has no registration history.")
            no_data.setStyleSheet("color: #6b7280; font-size: 13px;")
            layout.addWidget(no_data)

        dlg.exec()

    def load_event_combo(self):
        self.eventCombo.clear()
        events = Events()
        events.import_json("datasets/events.json")
        for e in events.list:
            self.eventCombo.addItem(f"{e.EventName} - {e.EventDate}", e.EventId)

    def load_registrations(self):
        if self.eventCombo.count() == 0:
            self.registrationTable.setRowCount(0)
            return
        event_id = self.eventCombo.currentData()
        if not event_id:
            return
        regs = Registrations()
        regs.import_json("datasets/registrations.json")
        attendees = Attendees()
        attendees.import_json("datasets/attendees.json")
        items = regs.get_registrations_by_event(event_id)
        self._reg_attendee_map = {}
        for reg in items:
            att = attendees.find_attendee(reg.AttendeeId)
            if att:
                self._reg_attendee_map[reg.RegistrationId] = (reg, att)

        valid_items = [reg for reg in items if reg.RegistrationId in self._reg_attendee_map]
        self._setup_pagination(
            self.registrationTable, valid_items,
            self._fill_registration_table,
            self.lblPageRegistration, self.btnPrevRegistration, self.btnNextRegistration
        )

    def _fill_registration_table(self, reg_list):
        self.registrationTable.setRowCount(len(reg_list))
        for row, reg in enumerate(reg_list):
            pair = self._reg_attendee_map.get(reg.RegistrationId)
            if not pair:
                continue
            reg, att = pair
            is_checkedin = reg.Status == "Checked-in"
            values = [
                reg.RegistrationId, att.Name, att.Email,
                att.Organization or "", reg.RegistrationDate,
                "✅ " + reg.Status if is_checkedin else reg.Status,
            ]
            for col, val in enumerate(values):
                item = QTableWidgetItem(val)
                if is_checkedin:
                    item.setBackground(QtGui.QColor("#F9E79F"))
                    item.setForeground(QtGui.QColor("#7D6608"))
                self.registrationTable.setItem(row, col, item)
        self._pad_table(self.registrationTable, len(reg_list))

    def register_attendee(self):
        if self.eventCombo.count() == 0:
            _msg(self.MainWindow, "warn", "Error", "No events available!")
            return
        event_id = self.eventCombo.currentData()
        dlg = RegistrationDialogEx(self.MainWindow, event_id=event_id)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            att_ids = dlg.get_selected_attendee_ids()
            if not att_ids:
                return

            regs = Registrations()
            regs.import_json("datasets/registrations.json")

            success_list = []
            skip_list = []

            for att_id in att_ids:
                if regs.find_registration_by_event_attendee(event_id, att_id):
                    skip_list.append(att_id)
                    continue
                reg = Registration()
                reg.RegistrationId = str(uuid.uuid4())[:8].upper()
                reg.EventId = event_id
                reg.AttendeeId = att_id
                reg.RegistrationDate = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                reg.Status = "Registered"
                regs.add_item(reg)
                success_list.append(reg.RegistrationId)

            if success_list:
                regs.export_json("datasets/registrations.json")

            msg = ""
            if success_list:
                codes = ", ".join(success_list)
                msg += f"✅ Registered {len(success_list)} attendee(s)!\nCodes: {codes}"
            if skip_list:
                msg += f"\n⚠️ {len(skip_list)} attendee(s) already registered (skipped)."

            # ── Hỏi gửi email sau khi đăng ký thành công ──
            if msg:
                _msg(self.MainWindow, "info", "Registration Result", msg)

                if success_list:
                    reply = QMessageBox.question(
                        self.MainWindow, "Send Emails",
                        "Do you want to send QR Codes via email to the newly registered attendees?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )

                    if reply == QMessageBox.StandardButton.Yes:
                        events = Events()
                        events.import_json("datasets/events.json")
                        event_name = events.find_event(event_id).EventName

                        attendees = Attendees()
                        attendees.import_json("datasets/attendees.json")
                        email_data = []

                        for reg_id in success_list:
                            reg = regs.find_registration(reg_id)
                            att = attendees.find_attendee(reg.AttendeeId)
                            if att and att.Email:
                                email_data.append({
                                    'email': att.Email,
                                    'name': att.Name,
                                    'code': reg_id
                                })

                        if email_data:
                            self.progress_dialog = QProgressDialog(
                                "Starting...", "Cancel", 0, len(email_data), self.MainWindow
                            )
                            self.progress_dialog.setWindowTitle("Sending Emails")
                            self.progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
                            self.progress_dialog.setMinimumDuration(0)

                            self.email_thread = EmailSenderThread(email_data, event_name)

                            def update_progress(current, text):
                                self.progress_dialog.setValue(current)
                                self.progress_dialog.setLabelText(text)

                            def email_finished(success_count, fail_count):
                                self.progress_dialog.close()
                                _msg(self.MainWindow, "info", "Email Result",
                                     f"Emails sent successfully: {success_count}\nFailed: {fail_count}")

                            self.email_thread.progress.connect(update_progress)
                            self.email_thread.finished_task.connect(email_finished)
                            self.email_thread.start()

            self.load_registrations()

    def cancel_registration(self):
        rows = list(set(idx.row() for idx in self.registrationTable.selectedIndexes()))
        if not rows:
            QMessageBox.warning(self.MainWindow, "Error", "Please select at least one registration!")
            return
        reply = QMessageBox.question(
            self.MainWindow, "Confirm",
            f"Cancel {len(rows)} registration(s)?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            reg_ids = [self.registrationTable.item(r, 0).text() for r in rows]
            regs = Registrations()
            regs.import_json("datasets/registrations.json")
            for reg_id in reg_ids:
                regs.delete_registration(reg_id)
            regs.export_json("datasets/registrations.json")
            _msg(self.MainWindow, "info", "Success", f"Canceled {len(reg_ids)} registration(s) successfully!")
            self.load_registrations()

    def generate_qr_code(self):
        if not QR_AVAILABLE:
            _msg(self.MainWindow, "warn", "Error", "qrcode not installed!\nRun: pip install qrcode[pil]")
            return
        row = self.registrationTable.currentRow()
        if row < 0:
            _msg(self.MainWindow, "warn", "Error", "Please select a registration!")
            return

        from PyQt6.QtWidgets import QPushButton, QHBoxLayout, QFileDialog, QApplication
        reg_id = self.registrationTable.item(row, 0).text()
        att_name = self.registrationTable.item(row, 1).text()
        att_info = f"{self.registrationTable.item(row, 2).text()} · {self.registrationTable.item(row, 3).text()}"

        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(reg_id)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        pixmap = QPixmap.fromImage(QImage.fromData(buf.read()))

        dlg = QDialog(self.MainWindow)
        dlg.setWindowTitle("Check-in QR Code")
        dlg.setMinimumWidth(400)
        dlg.setStyleSheet("background:white;")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(20, 20, 20, 20)
        lay.setSpacing(10)

        img_lbl = QLabel()
        img_lbl.setPixmap(pixmap)
        img_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(img_lbl)

        info_lbl = QLabel(
            f"<div style='text-align:center'><h3 style='color:#3b82f6;margin:0'>{att_name}</h3>"
            f"<p style='color:#64748b;font-size:12px;margin:2px'>{att_info}</p>"
            f"<h2 style='color:#1e40af'>🎫 {reg_id}</h2></div>"
        )
        info_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info_lbl.setTextFormat(Qt.TextFormat.RichText)
        lay.addWidget(info_lbl)

        _btn_style = lambda bg: (f"QPushButton{{background:{bg};color:white;border-radius:6px;"
                                 f"font-size:13px;padding:8px 16px;font-weight:bold}}"
                                 f"QPushButton:hover{{background:#1e293b}}")
        btn_save = QPushButton("💾 Save Image")
        btn_save.setStyleSheet(_btn_style("#3b82f6"))
        btn_copy = QPushButton("📋 Copy Code")
        btn_copy.setStyleSheet(_btn_style("#2563eb"))
        btn_close = QPushButton("✖ Close")
        btn_close.setStyleSheet(_btn_style("#ef4444"))

        row_btns = QHBoxLayout()
        for b in (btn_save, btn_copy, btn_close):
            row_btns.addWidget(b)
        lay.addLayout(row_btns)

        def save_qr():
            path, _ = QFileDialog.getSaveFileName(dlg, "Save QR",
                                                  f"QR_{att_name.replace(' ', '_')}_{reg_id}.png", "PNG (*.png)")
            if path:
                buf2 = BytesIO()
                img.save(buf2, format="PNG")
                open(path, "wb").write(buf2.getvalue())
                QMessageBox.information(dlg, "Saved", f"Saved!\n{path}")

        btn_save.clicked.connect(save_qr)
        btn_copy.clicked.connect(lambda: (QApplication.clipboard().setText(reg_id),
                                          QMessageBox.information(dlg, "Copied", f"Copied: {reg_id}")))
        btn_close.clicked.connect(dlg.accept)
        dlg.exec()

    def load_checkin_event_combo(self):
        self.checkinEventCombo.clear()
        events = Events()
        events.import_json("datasets/events.json")
        for e in events.list:
            self.checkinEventCombo.addItem(f"{e.EventName} - {e.EventDate}", e.EventId)

    def load_checkin_stats(self):
        if self.checkinEventCombo.count() == 0:
            self.totalRegisteredLabel.setText("0")
            self.totalCheckedinLabel.setText("0")
            self.checkinTable.setRowCount(0)
            return
        event_id = self.checkinEventCombo.currentData()
        if not event_id:
            return
        regs = Registrations()
        regs.import_json("datasets/registrations.json")
        self.totalRegisteredLabel.setText(str(regs.count_registered_by_event(event_id)))
        self.totalCheckedinLabel.setText(str(regs.count_checkedin_by_event(event_id)))
        attendees = Attendees()
        attendees.import_json("datasets/attendees.json")
        checkedin = [r for r in regs.get_registrations_by_event(event_id)
                     if r.Status == "Checked-in" or r.Status == "Đã check-in"]
        self.checkinTable.setRowCount(len(checkedin))
        for row, reg in enumerate(checkedin):
            att = attendees.find_attendee(reg.AttendeeId)
            if att:
                self.checkinTable.setItem(row, 0, QTableWidgetItem(att.Name))
                self.checkinTable.setItem(row, 1, QTableWidgetItem(att.Email))
                self.checkinTable.setItem(row, 2, QTableWidgetItem(att.Organization or ""))
                self.checkinTable.setItem(row, 3, QTableWidgetItem(reg.CheckinTime or ""))
                self.checkinTable.setItem(row, 4, QTableWidgetItem(reg.RegistrationId))
        self._pad_table(self.checkinTable, len(checkedin))

    def perform_checkin(self):
        code = self.checkinCode.text().strip().upper()
        if not code:
            _msg(self.MainWindow, "warn", "Error", "Please enter a registration code!")
            return

        event_id = self.checkinEventCombo.currentData()
        if not event_id:
            _msg(self.MainWindow, "warn", "Error", "Please select an event first!")
            return

        regs = Registrations()
        regs.import_json("datasets/registrations.json")
        success, message = regs.checkin_for_event(code, event_id)
        if success:
            regs.export_json("datasets/registrations.json")
            _msg(self.MainWindow, "info", "✅ Check-in", message)
            self.checkinCode.clear()
            self.load_checkin_stats()
            self.load_dashboard()
        else:
            _msg(self.MainWindow, "warn", "Check-in Failed", message)

    def scan_qr_checkin(self):
        from Ui_ex.QRScannerDialogEx import QRScannerDialogEx

        event_id = self.checkinEventCombo.currentData()
        if not event_id:
            _msg(self.MainWindow, "warn", "Error", "Please select an event first!")
            return

        def checkin_callback(qr_code):
            regs = Registrations()
            regs.import_json("datasets/registrations.json")
            success, message = regs.checkin_for_event(qr_code.upper(), event_id)
            if success:
                regs.export_json("datasets/registrations.json")
                QTimer.singleShot(500, self.load_checkin_stats)
            return success, message

        scanner = QRScannerDialogEx(self.MainWindow, callback=checkin_callback)
        scanner.exec()

    def _get_file_path(self, default_filename):
        from PyQt6.QtWidgets import QFileDialog
        file_path, _ = QFileDialog.getSaveFileName(
            self.MainWindow, "Export to CSV", default_filename,
            "CSV Files (*.csv);;All Files (*)"
        )
        return file_path

    def _write_csv(self, file_path, headers, rows):
        import csv
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

    def _do_export(self, filename, headers, build_rows):
        path = self._get_file_path(filename)
        if not path: return
        try:
            rows = build_rows()
            self._write_csv(path, headers, rows)
            _msg(self.MainWindow, "info", "Exported", f"Exported {len(rows)} records!\n{path}")
        except Exception as e:
            _msg(self.MainWindow, "warn", "Error", f"Export failed:\n{e}")

    def export_events_csv(self):
        events = Events()
        events.import_json("datasets/events.json")
        self._do_export("events.csv",
                        ["ID", "Event Name", "Date", "Time", "Venue", "Description"],
                        lambda: [[e.EventId, e.EventName, e.EventDate, e.EventTime,
                                  e.Location, e.Description or ""] for e in events.list])

    def export_attendees_csv(self):
        atts = Attendees()
        atts.import_json("datasets/attendees.json")
        self._do_export("attendees.csv",
                        ["ID", "Full Name", "Email", "Phone", "Organization", "Position"],
                        lambda: [[a.AttendeeId, a.Name, a.Email, a.Phone or "",
                                  a.Organization or "", a.Position or ""] for a in atts.list])

    def export_registrations_csv(self):
        regs = Registrations()
        regs.import_json("datasets/registrations.json")
        atts = Attendees()
        atts.import_json("datasets/attendees.json")
        event_id = self.eventCombo.currentData()
        items = regs.get_registrations_by_event(event_id) if event_id else regs.list

        def _rows():
            out = []
            for r in items:
                a = atts.find_attendee(r.AttendeeId)
                out.append([r.RegistrationId, a.Name if a else "", a.Email if a else "",
                            (a.Organization or "") if a else "",
                            r.RegistrationDate, r.Status])
            return out

        self._do_export("registrations.csv",
                        ["Reg. ID", "Full Name", "Email", "Organization", "Reg. Date", "Status"], _rows)

    def export_checkin_csv(self):
        regs = Registrations()
        regs.import_json("datasets/registrations.json")
        atts = Attendees()
        atts.import_json("datasets/attendees.json")
        event_id = self.checkinEventCombo.currentData()
        items = [r for r in (regs.get_registrations_by_event(event_id) if event_id else regs.list)
                 if r.Status == "Checked-in"]

        def _rows():
            out = []
            for r in items:
                a = atts.find_attendee(r.AttendeeId)
                out.append([a.Name if a else "", a.Email if a else "",
                            (a.Organization or "") if a else "",
                            getattr(r, "CheckinTime", ""), r.RegistrationId])
            return out

        self._do_export("checkin_list.csv",
                        ["Full Name", "Email", "Organization", "Check-in Time", "Reg. Code"], _rows)

    def _setup_pagination(self, table, data_list, fill_func, page_label, prev_btn, next_btn):
        if not hasattr(self, '_pages'):
            self._pages = {}

        name = table.objectName()
        self._pages[name] = {
            'data': data_list,
            'page': 0,
            'fill_func': fill_func,
            'label': page_label,
            'prev': prev_btn,
            'next': next_btn,
        }
        self._render_page(name)

    def _render_page(self, name):
        state = self._pages[name]
        data = state['data']
        page = state['page']
        total = max(1, (len(data) + self.PAGE_SIZE - 1) // self.PAGE_SIZE)
        start = page * self.PAGE_SIZE
        end = start + self.PAGE_SIZE
        chunk = data[start:end]

        state['fill_func'](chunk)
        state['label'].setText(f"Page {page + 1} / {total}")
        state['prev'].setEnabled(page > 0)
        state['next'].setEnabled(page < total - 1)

    def _prev_page(self, name):
        if self._pages[name]['page'] > 0:
            self._pages[name]['page'] -= 1
            self._render_page(name)

    def _next_page(self, name):
        state = self._pages[name]
        total = max(1, (len(state['data']) + self.PAGE_SIZE - 1) // self.PAGE_SIZE)
        if state['page'] < total - 1:
            state['page'] += 1
            self._render_page(name)

    def setup_chart(self):
        if not MATPLOTLIB_AVAILABLE:
            lbl = QLabel("⚠️ matplotlib not installed!\npip install matplotlib")
            lbl.setStyleSheet("color: red; font-size: 13px; padding: 20px;")
            self.verticalLayoutStatsPlot.addWidget(lbl)
            return

        self.stats_figure, self.stats_ax = plt.subplots()
        self.stats_figure.patch.set_facecolor("white")
        self.stats_canvas = FigureCanvas(self.stats_figure)

        self.verticalLayoutStatsPlot.addWidget(self.stats_canvas)
        self.show_bar_chart()

    def _set_active_btn(self, active_btn):
        for btn in [self.btnStatsBar, self.btnStatsLine, self.btnStatsPie]:
            if btn is active_btn:
                btn.setStyleSheet(btn.property("active_style"))
            else:
                btn.setStyleSheet(btn.property("inactive_style"))

    def load_stats_event_combo(self):
        self.statsEventCombo.clear()
        self.statsEventCombo.addItem("🌐 All Events", None)
        events = Events()
        events.import_json("datasets/events.json")
        for e in events.list:
            self.statsEventCombo.addItem(f"{e.EventName} - {e.EventDate}", e.EventId)

    def _get_stats_data(self):
        event_id = self.statsEventCombo.currentData()
        events = Events()
        events.import_json("datasets/events.json")
        regs = Registrations()
        regs.import_json("datasets/registrations.json")

        data = {}
        target_events = events.list if event_id is None else [
            e for e in events.list if e.EventId == event_id
        ]

        for e in target_events:
            event_regs = regs.get_registrations_by_event(e.EventId)
            registered = len(event_regs)
            checkedin = sum(1 for r in event_regs if r.Status == "Checked-in")
            label = e.EventName if len(e.EventName) <= 18 else e.EventName[:15] + "..."
            data[label] = {"registered": registered, "checkedin": checkedin}

        return data

    def refresh_chart(self):
        if not MATPLOTLIB_AVAILABLE:
            return
        if not hasattr(self, 'stats_ax') or not hasattr(self, 'stats_canvas'):
            return
        title = self.stats_ax.get_title()
        if "Line" in title:
            self.show_line_chart()
        elif "Pie" in title:
            self.show_pie_chart()
        else:
            self.show_bar_chart()

    def _chart_init(self, btn_attr):
        if not MATPLOTLIB_AVAILABLE: return None
        if hasattr(self, btn_attr): self._set_active_btn(getattr(self, btn_attr))
        self.stats_figure.clear()
        ax = self.stats_figure.add_subplot(111)
        ax.set_facecolor("#f9f9f9")
        self.stats_ax = ax
        return ax

    def _chart_done(self):
        self.stats_figure.tight_layout()
        self.stats_canvas.draw()

    def show_bar_chart(self):
        ax = self._chart_init('btnStatsBar')
        if ax is None: return
        data = self._get_stats_data()
        if not data: return

        labels = list(data.keys())
        reg = [data[k]["registered"] for k in labels]
        chk = [data[k]["checkedin"] for k in labels]
        x, w = range(len(labels)), 0.35

        for bars, vals, color, lbl in [
            ([i - w / 2 for i in x], reg, "#3498db", "Registered"),
            ([i + w / 2 for i in x], chk, "#27ae60", "Checked-in"),
        ]:
            b = ax.bar(bars, vals, w, label=lbl, color=color, alpha=0.85)
            for bar in b:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.1,
                        str(int(bar.get_height())), ha='center', va='bottom', fontsize=9)

        ax.set_xticks(list(x))
        ax.set_xticklabels(labels, rotation=15, ha='right', fontsize=9)
        ax.set_ylabel("Number of People")
        ax.set_title("Registration vs Check-in by Event", fontsize=12, fontweight='bold')
        ax.legend()
        ax.grid(axis='y', alpha=0.4)
        self._chart_done()

    def show_line_chart(self):
        ax = self._chart_init('btnStatsLine')
        if ax is None: return
        from collections import Counter
        regs = Registrations()
        regs.import_json("datasets/registrations.json")
        event_id = self.statsEventCombo.currentData()
        items = regs.get_registrations_by_event(event_id) if event_id else regs.list
        if not items:
            _msg(self.MainWindow, "info", "Info", "No registration data!")
            return

        counts = Counter(r.RegistrationDate for r in items)
        dates = sorted(counts.keys())
        total = 0
        cumulative = []
        for d in dates:
            total += counts[d]
            cumulative.append(total)

        ax.plot(dates, cumulative, marker='o', color='#e67e22',
                linewidth=2, markersize=5, label="Cumulative Registrations")
        ax.fill_between(range(len(dates)), cumulative, alpha=0.15, color='#e67e22')

        n = len(cumulative)
        show_idx = {0, n - 1} | set(range(0, n, max(1, n // 6))) if n else set()
        for i in show_idx:
            ax.annotate(str(cumulative[i]), (dates[i], cumulative[i]),
                        textcoords="offset points", xytext=(0, 8), ha='center', fontsize=8)

        step = max(1, n // 8)
        ax.set_xticks(range(0, n, step))
        ax.set_xticklabels([dates[i] for i in range(0, n, step)], rotation=30, ha='right', fontsize=8)
        ax.set_ylabel("Cumulative Registrations")
        ax.set_title("Line Chart: Registration Trend", fontsize=12, fontweight='bold')
        ax.legend()
        ax.grid(alpha=0.4)
        self._chart_done()

    def show_pie_chart(self):
        ax = self._chart_init('btnStatsPie')
        if ax is None: return
        regs = Registrations()
        regs.import_json("datasets/registrations.json")
        event_id = self.statsEventCombo.currentData()
        items = regs.get_registrations_by_event(event_id) if event_id else regs.list
        chk = sum(1 for r in items if r.Status == "Checked-in")
        if not items:
            _msg(self.MainWindow, "info", "Info", "No registration data!")
            return

        sizes = [chk, len(items) - chk]
        labels = [f"Checked-in ({chk})", f"Not Checked-in ({len(items) - chk})"]
        wedges, _, autotexts = ax.pie(
            sizes, labels=labels, colors=["#27ae60", "#e74c3c"],
            autopct='%1.1f%%', explode=(0.05, 0), startangle=90,
            textprops={'fontsize': 10}
        )
        for at in autotexts:
            at.set_fontweight('bold')
        ax.set_title("Pie Chart: Check-in Rate", fontsize=12, fontweight='bold')
        ax.legend(wedges, labels, loc="lower right", fontsize=9)
        self._chart_done()

    def apply_stylesheet(self):
        self.MainWindow.setStyleSheet("""
            /* ── Theme B: Slate Blue ── */
            QMainWindow { background-color: white; }

            QPushButton {
                background-color: #3b82f6; color: white;
                border: none; padding: 8px 15px;
                border-radius: 5px; font-size: 12px;
            }
            QPushButton:hover   { background-color: #2563eb; }
            QPushButton:pressed { background-color: #1d4ed8; }
            QPushButton:disabled { background-color: #cbd5e1; color: #94a3b8; }

            QTableWidget {
                background-color: white;
                alternate-background-color: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                gridline-color: #e2e8f0;
            }
            QTableWidget::item { color: #111827; padding: 4px; }
            QTableWidget::item:selected { background-color: #dbeafe; color: #1e40af; }
            QHeaderView::section {
                background-color: #f1f5f9; color: #374151;
                padding: 8px; border: none;
                border-bottom: 2px solid #e2e8f0;
                font-weight: bold; font-size: 12px;
            }

            QLineEdit, QTextEdit, QComboBox, QDateEdit, QTimeEdit {
                padding: 6px; border: 1px solid #d1d5db;
                border-radius: 5px; background-color: white; color: #111827;
            }
            QLineEdit:focus, QComboBox:focus { border: 1px solid #3b82f6; }
            QLineEdit:disabled { background-color: #f9fafb; color: #9ca3af; }

            QGroupBox {
                border: 1px solid #e5e7eb; border-radius: 6px;
                margin-top: 10px; font-weight: bold;
                padding: 15px; background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin; subcontrol-position: top left;
                padding: 4px 10px; background-color: #f1f5f9;
                color: #374151; border-radius: 4px;
            }

            QTabWidget::pane { border: 1px solid #e5e7eb; border-radius: 6px; background: white; }
            QTabBar::tab {
                background: #f3f4f6; color: #6b7280;
                padding: 8px 16px; border-radius: 5px 5px 0 0;
                font-size: 12px; margin-right: 2px;
            }
            QTabBar::tab:selected { background: white; color: #111827; font-weight: bold; border-top: 2px solid #3b82f6; }
            QTabBar::tab:hover:!selected { background: #e5e7eb; color: #374151; }

            QScrollBar:vertical { background: #f9fafb; width: 8px; border-radius: 4px; }
            QScrollBar::handle:vertical { background: #d1d5db; border-radius: 4px; min-height: 20px; }
            QScrollBar:horizontal { background: #f9fafb; height: 8px; border-radius: 4px; }
            QScrollBar::handle:horizontal { background: #d1d5db; border-radius: 4px; min-width: 20px; }

            QLabel { color: #111827; }
            QCheckBox { color: #111827; }
            QWidget#contentWidget { background-color: #f9fafb; }

            QMessageBox { background-color: white; }
            QMessageBox QLabel { color: #111827; font-size: 13px; min-width: 280px; }
            QMessageBox QPushButton {
                background-color: #3b82f6; color: white; border: none;
                padding: 7px 22px; border-radius: 5px;
                font-size: 12px; min-width: 80px;
            }
            QMessageBox QPushButton:hover   { background-color: #2563eb; }
            QMessageBox QPushButton:pressed { background-color: #1d4ed8; }

            QDialog { background-color: white; }
            QDialog QLabel { color: #111827; }
        """)